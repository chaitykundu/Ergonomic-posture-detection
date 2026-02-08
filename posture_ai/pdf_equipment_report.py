import os
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def generate_equipment_excel(equipment_list):
    """
    Generate an Excel (XLSX) report for ergonomic equipment recommendations.
    """

    os.makedirs("output", exist_ok=True)

    timestamp = time.strftime("%Y-%m-%d_%H-%M-%S")
    output_xlsx = f"output/equipment_report_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment Recommendations"

    headers = [
        "ID",
        "Equipment Name",
        "Description",
        "Priority",
        "Expected Improvement",
        "Source"
        #"Target Issue",
        #"Price Range",
        #"Why Recommended",
    ]

    ws.append(headers)

    # Header styling
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DCE6F1", fill_type="solid")

    # Data rows
    for idx, item in enumerate(equipment_list, start=1):
        ws.append([
            idx,
            item.get("name", ""),
            item.get("description", ""),
            item.get("priority", "").upper(),
            item.get("improvement_percentage", ""),
            item.get("source", ""),
            # item.get("target_issue", ""),
            # item.get("price_range", ""),
            # item.get("description", ""),
            # item.get("why_recommended", ""),
            
        ])

    # Auto column width
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    wb.save(output_xlsx)
    return output_xlsx
